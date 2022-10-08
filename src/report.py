import openpyxl
from datetime import datetime
import os
import shutil
import json

try:
  date = os.sys.argv[1]
except:
  print('Please provide date as argument')
  print('Example: 2022-10')
  exit()
if len(date) != 7:
  print('Please enter the correct format')
  print('Example: 2022-10')
  exit()

def main():
  for file in os.listdir('output'):
    if date in file:
      print(file)

      fileName = f'output/{date}.xlsx'
      if not os.path.exists(fileName):
        shutil.copy('report.xlsx', fileName)

      book = openpyxl.load_workbook(fileName)
      sheet = book.active

      row = 8
      flag = True
      bodys = []

      while flag:
        value = sheet.cell(row=row, column=1).value
        if value == None or value == '日付': flag = False
        else: row += 1

      with open(f'output/{file}', 'r') as f:
        data = json.load(f)

        sheet.cell(row=row, column=1).value = data.get('day')
        for body in data.get('body'):
          bodys.append(body.get('content'))
        sheet.cell(row=row, column=3).value = max(bodys, key=len)
        sheet.cell(row=row, column=11).value = data.get('classroom')
        sheet.cell(row=row, column=14).value = len(data.get('students'))

      book.save(fileName)

if __name__ == '__main__':
  main()